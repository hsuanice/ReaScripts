#!/usr/bin/env python3
"""
subtitle_to_clb.py
Convert subtitle / dialogue-list files to CLB JSON for hsuanice_Subtitle Bridge.lua.

Two-phase API:
  python3 subtitle_to_clb.py --inspect <filepath>
      Returns metadata for non-SRT formats (sheets, columns, sample rows,
      auto-detected IN/OUT columns, text candidates) so the Lua side can
      ask the user which sheet / which text column to use.
      For SRT (which has no choices) returns events directly with ready=true.

  python3 subtitle_to_clb.py --parse <filepath>
      [--sheet=N] [--in-col=N] [--out-col=N] [--text-col=N] [--fps=24]
      Returns CLB events JSON.

Supported input formats:
  .srt                Absolute time HH:MM:SS,mmm
  .csv / .tsv         Header optional; left-most timecode column = IN, next = OUT
  .xlsx / .xlsm       Multi-sheet Excel; same auto-detect rule per sheet

CLB events JSON shape (parse mode):
  {
    "format":  "SUBTITLE_SRT" | "SUBTITLE_CSV" | "SUBTITLE_TSV" | "SUBTITLE_XLSX",
    "title":   <basename without ext>,
    "fps":     <float>,
    "is_drop": <bool>,
    "events":  [
      {
        "event_num": "000001",
        "reel": "", "track": "", "edit_type": "C",
        "src_tc_in": "00:00:00:00", "src_tc_out": "00:00:00:00",
        "rec_tc_in": "HH:MM:SS:FF", "rec_tc_out": "HH:MM:SS:FF",
        "clip_name": <text>,
        "source_file": "", "scene": "", "take": ""
      }, ...
    ]
  }

On error outputs: { "error": "...", "traceback": "..." }

Dependencies:
  openpyxl     (only for .xlsx/.xlsm)

Version: 260426.1900
"""

import sys
import os
import re
import csv
import json
import argparse
import traceback
from pathlib import Path


# ---------------------------------------------------------------------------
# Time parsing
# ---------------------------------------------------------------------------

SRT_TIME_RE = re.compile(r'(\d+):(\d+):(\d+)[,.](\d+)')
TC_RE = re.compile(r'^\s*(\d+):(\d+):(\d+)[:;.](\d+)\s*$')

# Lenient regex used only for column auto-detection — accepts both TC and
# SRT-style time strings within a cell value.
TIME_LIKE_RE = re.compile(
    r'^\s*\d{1,2}:\d{1,2}:\d{1,2}[:;.,]\d{1,3}\s*$'
)


def fps_info(fps, is_drop=False):
    """Return (nominal, factor, is_drop) for a given FPS value.

    Supports common rates: 23.976, 24, 25, 29.97 NDF/DF, 30, 48, 50, 59.94, 60.
    'factor' is the wall-clock per nominal-frame multiplier
    (e.g. 23.976 → 1001/1000)."""
    f = float(fps)
    if abs(f - 23.976) < 0.01:
        return 24, 1001 / 1000, False
    if abs(f - 29.97) < 0.01:
        return 30, 1001 / 1000, bool(is_drop)
    if abs(f - 59.94) < 0.01:
        return 60, 1001 / 1000, bool(is_drop)
    nom = int(round(f))
    return nom, 1.0, False


def srt_time_to_seconds(s):
    m = SRT_TIME_RE.search(s)
    if not m:
        raise ValueError(f"Bad SRT timestamp: {s}")
    h, mn, sec, ms = m.groups()
    return int(h) * 3600 + int(mn) * 60 + int(sec) + int(ms) / 1000.0


def tc_to_seconds(tc, fps, is_drop=False):
    nom, factor, df = fps_info(fps, is_drop)
    m = TC_RE.match(str(tc))
    if not m:
        raise ValueError(f"Bad timecode: {tc!r}")
    h, mn, s, f = map(int, m.groups())
    if df:
        total_min = h * 60 + mn
        drop = 2 * (total_min - total_min // 10)
        frames = (h * 3600 + mn * 60 + s) * nom + f - drop
    else:
        frames = (h * 3600 + mn * 60 + s) * nom + f
    return frames * factor / nom


def seconds_to_tc(seconds, fps, is_drop=False):
    """Convert wall-clock seconds → 'HH:MM:SS:FF' timecode string."""
    nom, factor, df = fps_info(fps, is_drop)
    if seconds < 0:
        seconds = 0.0
    total_frames = int(round(seconds * nom / factor))
    if df:
        # 29.97 / 59.94 drop-frame conversion (SMPTE)
        drop = 2 if abs(fps - 29.97) < 0.01 else 4  # 4 for 59.94 DF
        frames_per_10min = nom * 60 * 10 - drop * 9
        frames_per_min = nom * 60 - drop
        d = total_frames // frames_per_10min
        m = total_frames % frames_per_10min
        if m > drop:
            total_frames += drop * 9 * d + drop * ((m - drop) // frames_per_min)
        else:
            total_frames += drop * 9 * d
    fr = int(total_frames % nom)
    s = int((total_frames // nom) % 60)
    mn = int((total_frames // (nom * 60)) % 60)
    h = int(total_frames // (nom * 3600))
    sep = ';' if df else ':'
    return f"{h:02d}:{mn:02d}:{s:02d}{sep}{fr:02d}"


# ---------------------------------------------------------------------------
# File reading helpers
# ---------------------------------------------------------------------------

def read_text_flexible(path):
    """Read a text file, trying common encodings (UTF-8 BOM, UTF-8, Big5, GB18030,
    CP950) before falling back to UTF-8 with replacement."""
    for enc in ('utf-8-sig', 'utf-8', 'big5', 'gb18030', 'cp950'):
        try:
            return Path(path).read_text(encoding=enc), enc
        except UnicodeDecodeError:
            continue
    return Path(path).read_text(encoding='utf-8', errors='replace'), 'utf-8(replace)'


def sniff_csv_dialect(text, sample_size=4096):
    """Best-effort CSV dialect detection."""
    sample = text[:sample_size]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        return csv.excel  # default to comma


# ---------------------------------------------------------------------------
# SRT
# ---------------------------------------------------------------------------

def parse_srt(path):
    """Returns list of (start_sec, end_sec, text)."""
    text, _enc = read_text_flexible(path)
    blocks = re.split(r'\r?\n\r?\n+', text.strip())
    entries = []
    for block in blocks:
        lines = [l for l in block.splitlines() if l.strip()]
        tc_line = next((l for l in lines if '-->' in l), None)
        if not tc_line:
            continue
        s_str, e_str = tc_line.split('-->')
        try:
            start = srt_time_to_seconds(s_str)
            end = srt_time_to_seconds(e_str)
        except ValueError:
            continue
        idx = lines.index(tc_line)
        content = ' | '.join(lines[idx + 1:]).strip()
        if end > start and content:
            entries.append((start, end, content))
    return entries


# ---------------------------------------------------------------------------
# CSV / TSV / XLSX raw row extraction
# ---------------------------------------------------------------------------

def read_csv_tsv_rows(path, ext):
    text, enc = read_text_flexible(path)
    if ext == '.tsv':
        rows = [line.split('\t') for line in text.splitlines() if line.strip() != '']
    else:
        dialect = sniff_csv_dialect(text)
        rows = list(csv.reader(text.splitlines(), dialect))
        rows = [r for r in rows if any((c or '').strip() for c in r)]
    return rows, enc


def read_xlsx_workbook(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed.\n"
            "Run: pip3 install openpyxl"
        )
    return load_workbook(str(path), data_only=True, read_only=True)


def xlsx_sheet_rows(ws, max_rows=None):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and i >= max_rows:
            break
        if row and any(c is not None and str(c).strip() != '' for c in row):
            rows.append(['' if c is None else c for c in row])
    return rows


# ---------------------------------------------------------------------------
# Column auto-detection (CSV/TSV/XLSX)
# ---------------------------------------------------------------------------

def detect_in_out_columns(rows):
    """Find IN/OUT columns by scanning data rows for time-like cells.

    Rule (matches user spec): for each data row with ≥ 2 timecode-shaped cells,
    the LEFT-MOST is IN and the next-left is OUT. Pick the most-frequent
    (in_col, out_col) pair across all data rows.

    Returns (in_col, out_col, header_present) — column indices are 0-based.
    """
    pair_counts = {}
    rows_with_tc = 0
    for r in rows:
        tc_cols = [i for i, c in enumerate(r)
                   if c is not None and TIME_LIKE_RE.match(str(c).strip() or '')]
        if len(tc_cols) >= 2:
            pair = (tc_cols[0], tc_cols[1])
            pair_counts[pair] = pair_counts.get(pair, 0) + 1
            rows_with_tc += 1

    if not pair_counts:
        return None, None, False

    # Most-frequent pair wins; ties broken by left-most IN column
    best_pair, best_count = max(
        pair_counts.items(),
        key=lambda kv: (kv[1], -kv[0][0])
    )
    in_col, out_col = best_pair

    # Heuristic: if first row has no time-like cell in those columns, it's a header
    header_present = False
    if rows:
        first = rows[0]
        if (in_col >= len(first) or
                not TIME_LIKE_RE.match(str(first[in_col] or '').strip())):
            header_present = True
    return in_col, out_col, header_present


def make_columns_meta(rows, in_col, out_col, header_present):
    """Build columns-list and text-candidate column indices.

    columns: [{ index, name }]  (index 1-based for the Lua side)
    text_candidates: list of column indices (1-based) that look like text
                     (non-empty, non-timecode in at least one sample row)
    """
    if not rows:
        return [], []
    width = max(len(r) for r in rows)

    if header_present:
        header = rows[0]
        data_rows = rows[1:]
    else:
        header = []
        data_rows = rows

    columns = []
    for c in range(width):
        if c < len(header) and header[c] is not None and str(header[c]).strip() != '':
            name = str(header[c]).strip()
        else:
            name = f"Column {c + 1}"
        columns.append({"index": c + 1, "name": name})

    candidates = []
    for c in range(width):
        if c == in_col or c == out_col:
            continue
        # A column is a text candidate if at least one sample data row has
        # a non-empty cell that is not time-like
        for r in data_rows[:30]:
            if c >= len(r):
                continue
            v = ('' if r[c] is None else str(r[c])).strip()
            if v and not TIME_LIKE_RE.match(v):
                candidates.append(c + 1)
                break

    return columns, candidates


def sample_rows_serializable(rows, max_rows=8):
    """Convert raw rows to JSON-safe nested lists (truncated)."""
    out = []
    for r in rows[:max_rows]:
        out.append(['' if c is None else str(c) for c in r])
    return out


# ---------------------------------------------------------------------------
# Inspect mode
# ---------------------------------------------------------------------------

def inspect_file(path, default_fps=25.0):
    ext = Path(path).suffix.lower()

    if ext == '.srt':
        entries = parse_srt(path)
        events = [
            {
                "start_sec": s,
                "end_sec": e,
                "text": t,
            }
            for s, e, t in entries
        ]
        return {
            "format": "SRT",
            "ready": True,
            "events_seconds": events,
            "title": Path(path).stem,
        }

    if ext in ('.csv', '.tsv'):
        rows, enc = read_csv_tsv_rows(path, ext)
        in_col, out_col, header_present = detect_in_out_columns(rows)
        columns, text_candidates = make_columns_meta(
            rows, in_col, out_col, header_present)
        return {
            "format": "CSV" if ext == '.csv' else "TSV",
            "ready": False,
            "encoding": enc,
            "header_present": header_present,
            "columns": columns,
            "sample_rows": sample_rows_serializable(rows, max_rows=8),
            "detected_in_col": (in_col + 1) if in_col is not None else None,
            "detected_out_col": (out_col + 1) if out_col is not None else None,
            "text_candidates": text_candidates,
            "title": Path(path).stem,
        }

    if ext in ('.xlsx', '.xlsm'):
        wb = read_xlsx_workbook(path)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = xlsx_sheet_rows(ws, max_rows=200)
            in_col, out_col, header_present = detect_in_out_columns(rows)
            columns, text_candidates = make_columns_meta(
                rows, in_col, out_col, header_present)
            sheets.append({
                "name": sheet_name,
                "header_present": header_present,
                "columns": columns,
                "sample_rows": sample_rows_serializable(rows, max_rows=8),
                "detected_in_col": (in_col + 1) if in_col is not None else None,
                "detected_out_col": (out_col + 1) if out_col is not None else None,
                "text_candidates": text_candidates,
            })
        return {
            "format": "XLSX",
            "ready": False,
            "sheets": sheets,
            "title": Path(path).stem,
        }

    raise ValueError(f"Unsupported file type: {ext}")


# ---------------------------------------------------------------------------
# Parse mode
# ---------------------------------------------------------------------------

def parse_file(path, sheet=None, in_col=None, out_col=None, text_col=None,
               fps=25.0, is_drop=False):
    ext = Path(path).suffix.lower()

    entries = []  # list of (start_sec, end_sec, text)

    if ext == '.srt':
        entries = parse_srt(path)

    elif ext in ('.csv', '.tsv'):
        rows, _enc = read_csv_tsv_rows(path, ext)
        if in_col is None or out_col is None or text_col is None:
            d_in, d_out, header_present = detect_in_out_columns(rows)
            in_col = in_col or ((d_in + 1) if d_in is not None else None)
            out_col = out_col or ((d_out + 1) if d_out is not None else None)
        entries = rows_to_entries(rows, in_col, out_col, text_col, fps, is_drop)

    elif ext in ('.xlsx', '.xlsm'):
        wb = read_xlsx_workbook(path)
        names = wb.sheetnames
        if not names:
            raise ValueError("No sheets in workbook.")
        if sheet is None:
            sheet_name = names[0]
        else:
            # Try as name first (sheet names like "102" are valid and must
            # not be auto-coerced to indices). Fall back to 1-based index
            # only if no name matches.
            sheet_str = str(sheet)
            if sheet_str in names:
                sheet_name = sheet_str
            else:
                try:
                    idx = int(sheet_str)
                except ValueError:
                    raise ValueError(f"Sheet not found: {sheet!r}")
                if idx < 1 or idx > len(names):
                    raise ValueError(
                        f"Sheet not found: {sheet!r} "
                        f"(workbook has sheets: {names})"
                    )
                sheet_name = names[idx - 1]
        ws = wb[sheet_name]
        rows = xlsx_sheet_rows(ws)
        if in_col is None or out_col is None or text_col is None:
            d_in, d_out, header_present = detect_in_out_columns(rows)
            in_col = in_col or ((d_in + 1) if d_in is not None else None)
            out_col = out_col or ((d_out + 1) if d_out is not None else None)
        entries = rows_to_entries(rows, in_col, out_col, text_col, fps, is_drop)

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    # Build CLB events JSON
    events = []
    for i, (s, e, txt) in enumerate(entries, start=1):
        events.append({
            "event_num": f"{i:06d}",
            "reel": "",
            "track": "",
            "edit_type": "C",
            "src_tc_in": "00:00:00:00",
            "src_tc_out": "00:00:00:00",
            "rec_tc_in": seconds_to_tc(s, fps, is_drop),
            "rec_tc_out": seconds_to_tc(e, fps, is_drop),
            "clip_name": txt,
            "source_file": "",
            "scene": "",
            "take": "",
        })

    fmt_label = {
        '.srt':  "SUBTITLE_SRT",
        '.csv':  "SUBTITLE_CSV",
        '.tsv':  "SUBTITLE_TSV",
        '.xlsx': "SUBTITLE_XLSX",
        '.xlsm': "SUBTITLE_XLSX",
    }.get(ext, "SUBTITLE")

    return {
        "format":  fmt_label,
        "title":   Path(path).stem,
        "fps":     float(fps),
        "is_drop": bool(is_drop),
        "events":  events,
    }


def rows_to_entries(rows, in_col, out_col, text_col, fps, is_drop):
    """Convert tabular rows → list of (start_sec, end_sec, text).

    in_col / out_col / text_col are 1-based.
    Rows whose IN/OUT cells are not parseable timecodes are silently skipped
    (covers header rows and notes/comments at the top).
    """
    if not in_col or not out_col or not text_col:
        raise ValueError("rows_to_entries requires in_col, out_col, text_col")

    # Convert to 0-based
    ic, oc, tc = in_col - 1, out_col - 1, text_col - 1

    entries = []
    for r in rows:
        if max(ic, oc, tc) >= len(r):
            continue
        in_str = ('' if r[ic] is None else str(r[ic])).strip()
        out_str = ('' if r[oc] is None else str(r[oc])).strip()
        text = ('' if r[tc] is None else str(r[tc])).strip()
        if not in_str or not out_str or not text:
            continue
        if not TC_RE.match(in_str) or not TC_RE.match(out_str):
            continue
        try:
            s = tc_to_seconds(in_str, fps, is_drop)
            e = tc_to_seconds(out_str, fps, is_drop)
        except ValueError:
            continue
        if e <= s:
            continue
        entries.append((s, e, text))
    return entries


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_int_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == '':
        return None
    return int(s)


def main():
    parser = argparse.ArgumentParser(description="Subtitle to CLB JSON")
    parser.add_argument("filepath", help="Subtitle file path")
    parser.add_argument("--inspect", action="store_true",
                        help="Return file metadata instead of events")
    parser.add_argument("--parse", action="store_true",
                        help="Return CLB events JSON (default if --inspect not given)")
    parser.add_argument("--sheet",
                        help="Sheet name or 1-based index (xlsx only)")
    parser.add_argument("--in-col", type=int, default=None,
                        help="1-based IN column (csv/tsv/xlsx)")
    parser.add_argument("--out-col", type=int, default=None,
                        help="1-based OUT column (csv/tsv/xlsx)")
    parser.add_argument("--text-col", type=int, default=None,
                        help="1-based text column (csv/tsv/xlsx)")
    parser.add_argument("--fps", type=float, default=25.0,
                        help="Frame rate for TC conversion (default 25)")
    parser.add_argument("--drop-frame", action="store_true",
                        help="Treat fps as drop-frame")
    parser.add_argument("--progress-file", default=None,
                        help="Write progress lines to this path (unused for subtitles)")
    args = parser.parse_args()

    try:
        if not os.path.isfile(args.filepath):
            raise FileNotFoundError(f"File not found: {args.filepath}")

        # Sheet arg: pass through as string; parse_file looks up by name first,
        # then falls back to 1-based index, so numeric sheet names ("102") work.
        sheet = args.sheet if (args.sheet is not None and args.sheet != "") else None

        if args.inspect:
            data = inspect_file(args.filepath, default_fps=args.fps)
        else:
            data = parse_file(
                args.filepath,
                sheet=sheet,
                in_col=args.in_col,
                out_col=args.out_col,
                text_col=args.text_col,
                fps=args.fps,
                is_drop=args.drop_frame,
            )
        sys.stdout.write(json.dumps(data, ensure_ascii=False))
        sys.stdout.flush()
    except Exception as e:
        err = {
            "error": str(e),
            "traceback": traceback.format_exc(),
        }
        sys.stdout.write(json.dumps(err, ensure_ascii=False))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()
